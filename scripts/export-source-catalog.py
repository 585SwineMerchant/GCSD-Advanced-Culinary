#!/usr/bin/env python3
"""Export the Advanced Culinary intake workbook to a browser-readable snapshot."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def table_rows(workbook, sheet_name: str, header_row: int = 4) -> list[dict]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[header_row]]
    rows = []
    for cells in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None for value in cells):
            continue
        rows.append({key: value for key, value in zip(headers, cells) if key})
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=False, data_only=True)
    recipes = table_rows(workbook, "Recipe Bank")
    references = table_rows(workbook, "Quick References")

    payload = {
        "title": "Advanced Culinary Recipe Source Catalog",
        "source": args.workbook.name,
        "sourceImageCount": 177,
        "statusNote": (
            "These records identify complete source recipes and formulas. They remain source "
            "candidates until ingredients, procedure, equipment, allergens, and supplier matches "
            "are transcribed, checked, tested, and teacher-approved."
        ),
        "recipes": [
            {
                "id": row.get("Recipe ID"),
                "name": row.get("Recipe / Formula"),
                "category": row.get("Category"),
                "subcategory": row.get("Subcategory"),
                "sourceImages": row.get("Source Images"),
                "capturedYield": row.get("Captured Yield"),
                "courseRoute": row.get("Course Route"),
                "courseAlignment": row.get("Course Unit / Alignment"),
                "eventUses": row.get("Likely Event Uses"),
                "priority": row.get("Priority"),
                "captureType": row.get("Capture Type"),
                "databaseStatus": row.get("Database Status"),
                "notes": row.get("Notes"),
            }
            for row in recipes
        ],
        "references": [
            {
                "id": row.get("Reference ID"),
                "topic": row.get("Topic"),
                "type": row.get("Type"),
                "sourceImages": row.get("Source Images"),
                "coreIdea": row.get("Core Idea / Use"),
                "primaryCourse": row.get("Primary Course"),
                "placement": row.get("Suggested Placement"),
                "advancedFunction": row.get("Advanced / KM Function"),
                "priority": row.get("Priority"),
                "status": row.get("Status"),
            }
            for row in references
        ],
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
